from openpyxl import load_workbook
from allclass import *
def readresource(path):
    workbook = load_workbook(path)
    sheet=workbook['资源日历']
    id=0
    allresource=[]
    for row in sheet.iter_rows(values_only=True):
            # 检查行是否为空
        if(id==0):
            id+=1
            continue
        tem=resource()
        data=[]
        for id,col in enumerate(row):
            if(col!=None):
                data.append(col)
            elif(col==None and id==1):
                data.append('箔绕人员')
            elif(id>=8):
                continue
            else:
                break
        else:
            print(data)
            tem.initfromlist(data)
            allresource.append(tem)
    return allresource
    
def readquizfile(path):
    # 加载工作簿
    shengchangongdan=[{'工单编号':'', '物料编号':'', '数量':'', '计划开始日期':'', '计划完工日期':'', '前置工单':'', '优先级':''}]
    gongyiluxian=[{'物料编码':'', '工序':'', '前置工序':'', '资源名称':'', '资源需求':'', '准备工时min\n（固定值，资源占用）':'', '作业工时min\n（单位值，资源占用）':'', '后置工时min（固定值，不占用资源）':''}]
    ziyuanrili=[{'资源编号':'', '资源名称':'', '资源数量':'', '开始日期':'', '结束日期':'', '开始时间\n（包含）':'', '结束时间\n(包含)':'', '优先级':'', '班次':''}]
    gongdangongxu=[{'工单编号':'', '物料编号':'', '工序':'', '资源池':'', '资源ID':'', '资源需求':'', '开始时间':'', '结束时间':''}]
    gongxurenwu=[{}]   #?
    paichengjieguo=[{}]  #?
    workbook = load_workbook(path)

    # 获取所有工作表名
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        print(sheet_name)
    # 遍历每个工作表并读取数据
    for sheet_name in sheet_names:
        # 获取当前工作表
        sheet = workbook[sheet_name]

        # 打印工作表名
        print("表名:", sheet_name)

        # 遍历当前工作表的所有行
        for row in sheet.iter_rows(values_only=True):
            # 检查行是否为空
            if all(value is None for value in row):
                continue  # 如果行为空，则跳过
            # 打印非空行数据
            print(row)
    return shengchangongdan,gongyiluxian,ziyuanrili,gongdangongxu,gongxurenwu,paichengjieguo

# 用法示例
readresource('data.xlsx')
